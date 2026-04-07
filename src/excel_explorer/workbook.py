from pathlib import Path
from openpyxl import load_workbook as _openpyxl_load
from openpyxl.workbook import Workbook


def load_workbook(path: str, data_only: bool = False) -> Workbook:
    """Load an Excel workbook. Raises FileNotFoundError if path doesn't exist."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return _openpyxl_load(str(p), data_only=data_only)


def get_sheet_names(wb: Workbook) -> list[str]:
    """Return list of sheet names."""
    return wb.sheetnames


def get_cell_value(wb: Workbook, sheet: str, cell: str):
    """Get the value of a specific cell."""
    return wb[sheet][cell].value


def get_named_ranges(wb: Workbook) -> dict[str, str]:
    """Return dict of named range name -> target reference string."""
    result = {}
    for name, defn in wb.defined_names.items():
        result[name] = defn.attr_text
    return result
