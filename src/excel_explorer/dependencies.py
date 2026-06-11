import re
from pathlib import Path
from collections import defaultdict
from openpyxl.utils import get_column_letter
from excel_explorer.workbook import load_workbook as load_wb
from excel_explorer.formatters import format_output

import formulas as fml


def build_dag(path: str):
    """Load workbook into formulas ExcelModel and build the dependency graph."""
    model = fml.ExcelModel().loads(path)
    model.finish()
    return model


def _parse_node_ref(node_str: str, book_name: str) -> tuple[str, str] | None:
    """Parse a DAG node string into (sheet, cell) tuple.

    Node format: '[book.xlsx]SHEET'!A1 or '[book.xlsx]SHEET'!A1:B5
    Returns None for external file refs or non-cell nodes.
    """
    s = str(node_str)
    if "file:/" in s:
        return None

    # Match pattern: '[book]SHEET'!CELL
    pattern = rf"'\[{re.escape(book_name)}\]([^']+)'!([A-Z]+\d+)$"
    m = re.search(pattern, s, re.IGNORECASE)
    if m:
        return (m.group(1), m.group(2))
    return None


def _is_formula_node(node_str: str) -> bool:
    """Check if a node is a formula (computation) node vs a data node."""
    return str(node_str).startswith("=")


def _get_book_name(path: str) -> str:
    """Extract the workbook filename from path."""
    return Path(path).name


def _build_sheet_name_map(wb) -> dict[str, str]:
    """Build a mapping from uppercase sheet names to their original case."""
    return {name.upper(): name for name in wb.sheetnames}


def _find_cell_node(dmap, book_name: str, sheet: str, cell: str) -> str | None:
    """Find the DAG node matching a sheet!cell reference."""
    target = f"'[{book_name}]{sheet.upper()}'!{cell.upper()}"
    for node in dmap.nodes:
        if str(node).upper() == target.upper():
            return node
    # Anchor the cell ref so A1 doesn't match A10/A100 or ranges like A11:B5
    fallback = re.compile(rf"\]{re.escape(sheet.upper())}'!{re.escape(cell.upper())}(?=$|:)")
    for node in dmap.nodes:
        s = str(node).upper()
        if fallback.search(s) and "file:/" not in s and not s.startswith("="):
            return node
    return None


def _get_cell_value_safe(wb, sheet_name: str, cell: str, sheet_map: dict[str, str]):
    """Get cell value, mapping uppercase DAG sheet names back to original case."""
    original_name = sheet_map.get(sheet_name.upper(), sheet_name)
    try:
        return wb[original_name][cell].value
    except (KeyError, ValueError):
        return "?"


def trace_cell(path: str, sheet: str, cell: str, depth: int = 5) -> str:
    """Trace the full dependency tree from a cell down to leaf inputs."""
    model = build_dag(path)
    dmap = model.dsp.dmap
    book_name = _get_book_name(path)
    wb = load_wb(path)
    sheet_map = _build_sheet_name_map(wb)

    start_node = _find_cell_node(dmap, book_name, sheet, cell)

    if start_node is None:
        meta = {"file": book_name, "cell": f"{sheet}!{cell}", "depth": depth}
        return format_output(meta, f"Cell {sheet}!{cell} not found in dependency graph (may be a plain value with no dependents).")

    cell_value = _get_cell_value_safe(wb, sheet, cell, sheet_map)
    formula_str = cell_value if isinstance(cell_value, str) and cell_value.startswith("=") else str(cell_value)

    lines = []
    seen = set()
    shared_deps = defaultdict(int)
    leaf_inputs = []

    def _trace(node, prefix: str, is_last: bool, current_depth: int):
        ref = _parse_node_ref(str(node), book_name)
        node_label = f"{ref[0]}!{ref[1]}" if ref else str(node)[:60]

        if ref:
            cell_val = _get_cell_value_safe(wb, ref[0], ref[1], sheet_map)
        else:
            cell_val = ""

        is_formula = isinstance(cell_val, str) and cell_val.startswith("=")
        is_input = not is_formula and ref is not None

        display = f"{node_label} = {cell_val}"
        if is_input and ref:
            display += "  [INPUT]"
            leaf_inputs.append(f"{ref[0]}!{ref[1]}")

        if ref:
            key = f"{ref[0]}!{ref[1]}"
            shared_deps[key] += 1
            if shared_deps[key] > 1:
                display += " (shared)"

        connector = "\u2514\u2500\u2500 " if is_last else "\u251c\u2500\u2500 "
        lines.append(f"{prefix}{connector}{display}")

        if str(node) in seen or current_depth >= depth:
            if current_depth >= depth and not is_input:
                # Backslash escapes inside f-string expressions are a SyntaxError on 3.11
                bar = "\u2502   "
                lines.append(f"{prefix}{'    ' if is_last else bar}... (max depth)")
            return
        seen.add(str(node))

        child_prefix = prefix + ("    " if is_last else "\u2502   ")
        pred_nodes = []

        direct_preds = list(dmap.pred.get(node, {}).keys())
        for p in direct_preds:
            if _is_formula_node(str(p)):
                formula_preds = list(dmap.pred.get(p, {}).keys())
                pred_nodes.extend(formula_preds)
            else:
                pred_nodes.append(p)

        pred_nodes = [p for p in pred_nodes if "file:/" not in str(p) and not _is_formula_node(str(p))]

        for i, pred in enumerate(pred_nodes):
            is_last_child = (i == len(pred_nodes) - 1)
            _trace(pred, child_prefix, is_last_child, current_depth + 1)

    root_ref = _parse_node_ref(str(start_node), book_name)
    root_label = f"{root_ref[0]}!{root_ref[1]}" if root_ref else str(start_node)[:60]
    lines.append(f"{root_label} = {formula_str}")
    seen.add(str(start_node))

    direct_preds = list(dmap.pred.get(start_node, {}).keys())
    pred_nodes = []
    for p in direct_preds:
        if _is_formula_node(str(p)):
            formula_preds = list(dmap.pred.get(p, {}).keys())
            pred_nodes.extend(formula_preds)
        else:
            pred_nodes.append(p)
    pred_nodes = [p for p in pred_nodes if "file:/" not in str(p) and not _is_formula_node(str(p))]

    for i, pred in enumerate(pred_nodes):
        is_last = (i == len(pred_nodes) - 1)
        _trace(pred, "", is_last, 1)

    unique_inputs = list(set(leaf_inputs))
    shared = {k: v for k, v in shared_deps.items() if v > 1}
    if unique_inputs:
        lines.append(f"\nleaf_inputs: {len(unique_inputs)} unique")
        for inp in unique_inputs:
            lines.append(f"  - {inp}")
    if shared:
        lines.append(f"shared_deps:")
        for k, v in shared.items():
            lines.append(f"  - {k} (appears {v}x)")

    meta = {
        "file": book_name,
        "cell": f"{sheet}!{cell}",
        "formula": formula_str,
        "depth": depth,
    }
    return format_output(meta, "\n".join(lines))


def find_dependents(path: str, sheet: str, cell: str, depth: int = 5) -> str:
    """Find all cells that depend on this cell (reverse trace)."""
    model = build_dag(path)
    dmap = model.dsp.dmap
    book_name = _get_book_name(path)

    start_node = _find_cell_node(dmap, book_name, sheet, cell)
    if start_node is None:
        meta = {"file": book_name, "cell": f"{sheet}!{cell}"}
        return format_output(meta, f"Cell {sheet}!{cell} not found in dependency graph.")

    dependents = []
    seen = set()

    def _walk_succs(node, current_depth):
        if str(node) in seen or current_depth > depth:
            return
        seen.add(str(node))
        succs = list(dmap.succ.get(node, {}).keys())
        for s in succs:
            if _is_formula_node(str(s)):
                formula_succs = list(dmap.succ.get(s, {}).keys())
                for fs in formula_succs:
                    ref = _parse_node_ref(str(fs), book_name)
                    if ref:
                        dependents.append(f"{ref[0]}!{ref[1]}")
                    _walk_succs(fs, current_depth + 1)
            else:
                ref = _parse_node_ref(str(s), book_name)
                if ref:
                    dependents.append(f"{ref[0]}!{ref[1]}")
                _walk_succs(s, current_depth + 1)

    _walk_succs(start_node, 0)
    unique_deps = list(dict.fromkeys(dependents))

    lines = []
    lines.append(f"dependents of {sheet}!{cell}: {len(unique_deps)} cells")
    for d in unique_deps:
        lines.append(f"  - {d}")

    meta = {"file": book_name, "cell": f"{sheet}!{cell}", "depth": depth}
    return format_output(meta, "\n".join(lines))


def formula_map(path: str, sheet: str, limit: int = 50, offset: int = 0) -> str:
    """Show unique formula patterns on a sheet, grouping identical patterns."""
    wb = load_wb(path)
    ws = wb[sheet]
    book_name = _get_book_name(path)
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    patterns = defaultdict(list)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=min(max_col, 100)):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                normalized = re.sub(r'([A-Z]+)(\d+)', r'\1#', cell.value)
                patterns[normalized].append(cell.coordinate)

    sorted_patterns = sorted(patterns.items(), key=lambda x: -len(x[1]))
    total = len(sorted_patterns)
    sorted_patterns = sorted_patterns[offset:offset + limit]

    lines = []
    for pattern, cells in sorted_patterns:
        lines.append(f"  pattern: {pattern}")
        lines.append(f"    count: {len(cells)}")
        lines.append(f"    example_cells: {cells[:5]}")
        actual = ws[cells[0]].value
        lines.append(f"    example: {actual}")
        lines.append("")

    meta = {
        "file": book_name,
        "sheet": sheet,
        "unique_patterns": total,
        "showing": f"{offset + 1}-{offset + len(sorted_patterns)} of {total}",
        "truncated": offset + len(sorted_patterns) < total,
    }
    return format_output(meta, "\n".join(lines))


def find_inputs(path: str, sheet: str | None = None, limit: int = 50, offset: int = 0) -> str:
    """Find hardcoded cells that other cells depend on (model inputs/assumptions)."""
    model = build_dag(path)
    dmap = model.dsp.dmap
    book_name = _get_book_name(path)
    wb = load_wb(path)
    sheet_map = _build_sheet_name_map(wb)

    inputs = []
    for node in dmap.nodes:
        if _is_formula_node(str(node)) or "file:/" in str(node):
            continue
        ref = _parse_node_ref(str(node), book_name)
        if not ref:
            continue
        if sheet and ref[0].upper() != sheet.upper():
            continue

        succs = list(dmap.succ.get(node, {}).keys())
        if not succs:
            continue

        cell_val = _get_cell_value_safe(wb, ref[0], ref[1], sheet_map)
        if isinstance(cell_val, str) and cell_val.startswith("="):
            continue
        if cell_val is None:
            continue

        # Use original case sheet name for display
        original_sheet = sheet_map.get(ref[0].upper(), ref[0])
        inputs.append({
            "cell": f"{original_sheet}!{ref[1]}",
            "value": cell_val,
            "dependent_count": len(succs),
        })

    inputs.sort(key=lambda x: -x["dependent_count"])
    total = len(inputs)
    inputs = inputs[offset:offset + limit]

    lines = []
    for inp in inputs:
        lines.append(f"  {inp['cell']}: {inp['value']}  (feeds {inp['dependent_count']} nodes)")

    meta = {
        "file": book_name,
        "inputs_found": total,
        "showing": f"{offset + 1}-{offset + len(inputs)} of {total}",
        "truncated": offset + len(inputs) < total,
    }
    if sheet:
        meta["sheet"] = sheet
    return format_output(meta, "\n".join(lines))


def sheet_flow(path: str) -> str:
    """Map which sheets reference which other sheets."""
    wb = load_wb(path)
    book_name = _get_book_name(path)
    sheets = wb.sheetnames

    edges = set()
    all_referenced = set()
    all_referencing = set()

    for sn in sheets:
        ws = wb[sn]
        if not hasattr(ws, "iter_rows"):  # chartsheets have no cells
            continue
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 10000), max_col=min(max_col, 100)):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    refs = re.findall(r"'([^']+)'!", cell.value)
                    # \b instead of lookbehinds: (?<!=) made =Sheet!A1 match as
                    # a truncated 'heet', dropping refs at the start of a formula
                    refs += re.findall(r"\b(\w+)!", cell.value)
                    for ref in refs:
                        ref_clean = ref.strip("'")
                        if ref_clean != sn and ref_clean in sheets:
                            edges.add((ref_clean, sn))
                            all_referenced.add(ref_clean)
                            all_referencing.add(sn)

    targets = {e[1] for e in edges}
    origins = {e[0] for e in edges}
    sources = [s for s in sheets if s in origins and s not in targets]
    sinks = [s for s in sheets if s in targets and s not in origins]
    isolated = [s for s in sheets if s not in origins and s not in targets]

    lines = []
    if edges:
        lines.append("sheet_flow:")
        for src, dst in sorted(edges):
            lines.append(f'  "{src}" -> "{dst}"')

    if sources:
        lines.append("")
        lines.append("sources (no inbound refs):")
        for s in sources:
            lines.append(f"  - {s}")

    if sinks:
        lines.append("")
        lines.append("sinks (no outbound refs):")
        for s in sinks:
            lines.append(f"  - {s}")

    if isolated:
        lines.append("")
        lines.append("isolated (no cross-sheet refs):")
        for s in isolated:
            lines.append(f"  - {s}")

    meta = {
        "file": book_name,
        "sheets": len(sheets),
        "cross_sheet_edges": len(edges),
    }
    return format_output(meta, "\n".join(lines))
