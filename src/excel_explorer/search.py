from openpyxl.utils import get_column_letter

from excel_explorer.workbook import load_workbook
from excel_explorer.formatters import format_output

MAX_COLS = 100


def search(path: str, query: str, formulas: bool = False, limit: int = 50, offset: int = 0) -> str:
    """Search cells by value or formula content across all sheets.

    Case-insensitive. When formulas=True, also searches inside formula strings.
    Supports pagination via limit and offset.
    """
    wb = load_workbook(path, data_only=False)
    query_lower = query.lower()
    matches: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(max_col=MAX_COLS):
            for cell in row:
                val = cell.value
                if val is None:
                    continue

                matched = False
                col_letter = get_column_letter(cell.column)
                cell_ref = f"{sheet_name}!{col_letter}{cell.row}"

                # Check string/formula values
                if isinstance(val, str):
                    if formulas and val.startswith("="):
                        if query_lower in val.lower():
                            matched = True
                    elif not val.startswith("="):
                        if query_lower in val.lower():
                            matched = True
                    elif val.startswith("=") and not formulas:
                        # formula cell but formulas=False: skip formula content
                        pass
                else:
                    # numeric or other types: compare string representation
                    if query_lower in str(val).lower():
                        matched = True

                if matched:
                    matches.append(f"  {cell_ref}: {val!r}")

    total = len(matches)
    paginated = matches[offset: offset + limit]
    truncated = total > offset + limit

    lines = paginated if paginated else ["  (no matches)"]
    if truncated:
        remaining = total - (offset + limit)
        lines = lines + [f"  ... truncated ({remaining} more result(s))"]

    meta = {
        "command": "search",
        "file": path,
        "query": query,
        "formulas": formulas,
        "matches": total,
        "shown": len(paginated),
        "offset": offset,
        "limit": limit,
    }
    return format_output(meta, "\n".join(lines))


def find_formatting(path: str, sheet: str, limit: int = 50, offset: int = 0) -> str:
    """Report cells with notable formatting: bold, fill colors, borders.

    Checks:
    - Bold font
    - Non-default fill color (fgColor.rgb != '00000000')
    - Top or bottom border with a style set
    """
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]

    notable: list[str] = []

    for row in ws.iter_rows(max_col=MAX_COLS):
        for cell in row:
            if cell.value is None and cell.font is None and cell.fill is None:
                continue

            traits: list[str] = []

            # Bold check
            if cell.font and cell.font.bold:
                traits.append("bold")

            # Fill color check
            try:
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != "00000000":
                    traits.append(f"fill:{cell.fill.fgColor.rgb}")
            except Exception:
                pass

            # Border check (top or bottom)
            try:
                if cell.border:
                    if cell.border.top and cell.border.top.style:
                        traits.append(f"border-top:{cell.border.top.style}")
                    if cell.border.bottom and cell.border.bottom.style:
                        traits.append(f"border-bottom:{cell.border.bottom.style}")
            except Exception:
                pass

            if traits:
                col_letter = get_column_letter(cell.column)
                cell_ref = f"{col_letter}{cell.row}"
                val_repr = repr(cell.value) if cell.value is not None else "(empty)"
                notable.append(f"  {cell_ref}: {val_repr}  [{', '.join(traits)}]")

    total = len(notable)
    paginated = notable[offset: offset + limit]
    truncated = total > offset + limit

    lines = paginated if paginated else ["  (no notable formatting found)"]
    if truncated:
        remaining = total - (offset + limit)
        lines = lines + [f"  ... truncated ({remaining} more)"]

    meta = {
        "command": "find-formatting",
        "file": path,
        "sheet": sheet,
        "cells_with_formatting": total,
        "shown": len(paginated),
        "offset": offset,
        "limit": limit,
    }
    return format_output(meta, "\n".join(lines))
