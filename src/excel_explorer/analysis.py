import datetime
import re
from collections import defaultdict

from openpyxl.utils import get_column_letter

from excel_explorer.workbook import load_workbook, get_named_ranges
from excel_explorer.formatters import format_output


# ---------------------------------------------------------------------------
# summarize_assumptions
# ---------------------------------------------------------------------------

def summarize_assumptions(path: str, limit: int = 50, offset: int = 0) -> str:
    """Find named ranges and scan for assumption-style sheets, extracting
    hardcoded values with their labels."""
    wb = load_workbook(path)
    named = get_named_ranges(wb)

    assumption_sheet_names = {"assumptions", "inputs", "parameters", "drivers"}
    found_sheets = [s for s in wb.sheetnames if s.lower() in assumption_sheet_names]

    lines: list[str] = []

    # Named ranges section
    if named:
        lines.append("Named Ranges:")
        for name, ref in named.items():
            lines.append(f"  {name} = {ref}")
    else:
        lines.append("Named Ranges: none")

    # Assumption sheets section
    for sheet_name in found_sheets:
        lines.append(f"\nSheet: {sheet_name}")
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Skip header row if first row looks like a header
        start = 0
        if rows and all(isinstance(v, str) or v is None for v in rows[0]):
            start = 1
        entries = rows[start:]
        paginated = entries[offset: offset + limit]
        for row in paginated:
            if not any(v is not None for v in row):
                continue
            label = row[0] if row else ""
            value = row[1] if len(row) > 1 else ""
            if label is not None or value is not None:
                lines.append(f"  {label}: {value}")

    if not found_sheets and not named:
        lines.append("No assumption sheets or named ranges found.")

    meta = {
        "command": "summarize-assumptions",
        "path": path,
        "assumption_sheets": ", ".join(found_sheets) if found_sheets else "none",
        "named_ranges": len(named),
    }
    return format_output(meta, "\n".join(lines))


# ---------------------------------------------------------------------------
# compare_periods
# ---------------------------------------------------------------------------

def compare_periods(
    path: str, sheet: str, row: int, max_cols: int = 20, col_offset: int = 0
) -> str:
    """Read a time-series row, auto-detect date headers, compute
    period-over-period changes and growth rates, and summary stats."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    # Read header row (row 1) for period labels
    header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

    # Read the target row
    data_row = list(ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]

    # Determine label (column A)
    label = data_row[0] if data_row else f"Row {row}"

    # Collect numeric data starting from column 2 (index 1), respecting col_offset
    start_idx = 1 + col_offset
    end_idx = min(len(data_row), start_idx + max_cols)

    periods: list[str] = []
    values: list[float] = []

    for i in range(start_idx, end_idx):
        if i >= len(data_row):
            break
        val = data_row[i]
        if val is None:
            continue
        if not isinstance(val, (int, float)):
            continue

        # Build period label from header
        if i < len(header_row) and header_row[i] is not None:
            hdr = header_row[i]
            if isinstance(hdr, (datetime.datetime, datetime.date)):
                period_label = hdr.strftime("%Y-%m")
            else:
                period_label = str(hdr)
        else:
            period_label = get_column_letter(i + 1)

        periods.append(period_label)
        values.append(float(val))

    if not values:
        meta = {"command": "compare-periods", "path": path, "sheet": sheet, "row": row}
        return format_output(meta, f"No numeric data found in row {row}.")

    # Compute period-over-period changes and growth rates
    lines: list[str] = [f"Label: {label}", ""]
    lines.append(f"{'Period':<12}  {'Value':>12}  {'Change':>12}  {'Growth %':>10}")
    lines.append("-" * 52)

    for i, (period, val) in enumerate(zip(periods, values)):
        if i == 0:
            lines.append(f"{period:<12}  {val:>12,.2f}  {'—':>12}  {'—':>10}")
        else:
            change = val - values[i - 1]
            growth = (change / values[i - 1] * 100) if values[i - 1] != 0 else float("nan")
            lines.append(
                f"{period:<12}  {val:>12,.2f}  {change:>+12,.2f}  {growth:>9.1f}%"
            )

    # Summary stats
    lines.append("")
    lines.append("Summary:")
    lines.append(f"  Min:     {min(values):,.2f}  ({periods[values.index(min(values))]})")
    lines.append(f"  Max:     {max(values):,.2f}  ({periods[values.index(max(values))]})")
    lines.append(f"  Average: {sum(values)/len(values):,.2f}")
    if len(values) >= 2:
        total_growth = (values[-1] - values[0]) / values[0] * 100 if values[0] != 0 else float("nan")
        lines.append(f"  Total growth ({periods[0]} to {periods[-1]}): {total_growth:+.1f}%")

    meta = {
        "command": "compare-periods",
        "path": path,
        "sheet": sheet,
        "row": row,
        "label": label,
        "periods": len(values),
    }
    return format_output(meta, "\n".join(lines))


# ---------------------------------------------------------------------------
# find_anomalies
# ---------------------------------------------------------------------------

def _normalize_formula(val: str) -> str:
    """Replace row numbers with # to get a structural pattern."""
    return re.sub(r'([A-Z]+)(\d+)', r'\1#', val)


def find_anomalies(path: str, sheet: str, limit: int = 50, offset: int = 0) -> str:
    """Scan rows for formula pattern breaks: cells that break the dominant
    formula pattern of their row (different formula or hardcoded when others
    use formulas)."""
    wb_formulas = load_workbook(path, data_only=False)
    ws = wb_formulas[sheet]

    anomalies: list[str] = []

    rows = list(ws.iter_rows())
    paginated_rows = rows[offset: offset + limit]

    for row_cells in paginated_rows:
        row_num = row_cells[0].row

        # Skip header rows or rows with fewer than 3 data cells
        data_cells = [c for c in row_cells if c.column > 1 and c.value is not None]
        if len(data_cells) < 2:
            continue

        # Classify each cell as formula or literal and record pattern
        patterns: list[tuple] = []  # (col_index, col_letter, raw_value, pattern)
        for cell in data_cells:
            val = cell.value
            col_letter = get_column_letter(cell.column)
            if isinstance(val, str) and val.startswith("="):
                pattern = _normalize_formula(val)
                patterns.append((cell.column, col_letter, val, ("formula", pattern)))
            else:
                patterns.append((cell.column, col_letter, val, ("literal", None)))

        if not patterns:
            continue

        # Find the dominant type (formula vs literal)
        formula_count = sum(1 for _, _, _, p in patterns if p[0] == "formula")
        literal_count = len(patterns) - formula_count

        # Only analyze rows where there's a mix or a clear formula pattern
        if formula_count == 0:
            # All literals — no anomaly to detect
            continue

        # Find dominant formula pattern among formulas
        formula_patterns: list[str] = [p[1] for _, _, _, p in patterns if p[0] == "formula" and p[1] is not None]
        pattern_counts: dict[str, int] = defaultdict(int)
        for fp in formula_patterns:
            pattern_counts[fp] += 1

        dominant_pattern = max(pattern_counts, key=lambda k: pattern_counts[k]) if pattern_counts else None
        dominant_count = pattern_counts[dominant_pattern] if dominant_pattern else 0

        for col_idx, col_letter, val, (kind, pat) in patterns:
            cell_ref = f"{col_letter}{row_num}"

            if kind == "literal" and formula_count > 0:
                # Hardcoded value in a row that's mostly formulas
                anomalies.append(
                    f"{cell_ref}: hardcoded value ({val!r}) in formula row "
                    f"({formula_count} formula(s), {literal_count} literal(s))"
                )
            elif kind == "formula" and dominant_pattern and pat != dominant_pattern and dominant_count > 1:
                # Formula deviates from the dominant pattern
                anomalies.append(
                    f"{cell_ref}: formula pattern differs — got `{pat}`, "
                    f"dominant is `{dominant_pattern}`"
                )

    meta = {
        "command": "find-anomalies",
        "path": path,
        "sheet": sheet,
        "anomalies_found": len(anomalies),
    }

    if anomalies:
        body = f"Found {len(anomalies)} anomaly/anomalies:\n\n" + "\n".join(f"  • {a}" for a in anomalies)
    else:
        body = "No anomalies detected."

    return format_output(meta, body)


# ---------------------------------------------------------------------------
# validate_balance
# ---------------------------------------------------------------------------

_ASSET_KEYWORDS = {"total assets", "assets"}
_LIABILITY_KEYWORDS = {"total liabilities", "liabilities"}
_EQUITY_KEYWORDS = {"total equity", "equity", "stockholders equity", "shareholders equity"}


def _match_keyword(label: str, keyword_set: set[str]) -> bool:
    if label is None:
        return False
    return label.strip().lower() in keyword_set


def validate_balance(path: str, sheet: str) -> str:
    """Check that Assets = Liabilities + Equity across all period columns."""
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        meta = {"command": "validate-balance", "path": path, "sheet": sheet}
        return format_output(meta, "Sheet is empty.")

    # Find header row (row 0 / index 0) for period labels
    header_row = rows[0]

    assets_row: list | None = None
    liabilities_row: list | None = None
    equity_row: list | None = None

    for row in rows[1:]:
        if not row:
            continue
        label = row[0]
        if assets_row is None and _match_keyword(label, _ASSET_KEYWORDS):
            assets_row = list(row)
        elif liabilities_row is None and _match_keyword(label, _LIABILITY_KEYWORDS):
            liabilities_row = list(row)
        elif equity_row is None and _match_keyword(label, _EQUITY_KEYWORDS):
            equity_row = list(row)

    if assets_row is None or liabilities_row is None or equity_row is None:
        missing = []
        if assets_row is None:
            missing.append("Assets")
        if liabilities_row is None:
            missing.append("Liabilities")
        if equity_row is None:
            missing.append("Equity")
        meta = {"command": "validate-balance", "path": path, "sheet": sheet, "result": "ERROR"}
        return format_output(meta, f"Could not find rows for: {', '.join(missing)}.")

    # Check across all period columns (col index 1+)
    num_cols = max(len(assets_row), len(liabilities_row), len(equity_row))
    results: list[str] = []
    all_pass = True
    period_results: list[str] = []

    for col_idx in range(1, num_cols):
        a = assets_row[col_idx] if col_idx < len(assets_row) else None
        l = liabilities_row[col_idx] if col_idx < len(liabilities_row) else None
        e = equity_row[col_idx] if col_idx < len(equity_row) else None

        if a is None and l is None and e is None:
            continue

        # Build period label from header
        period_label: str
        if col_idx < len(header_row) and header_row[col_idx] is not None:
            hdr = header_row[col_idx]
            if isinstance(hdr, (datetime.datetime, datetime.date)):
                period_label = hdr.strftime("%Y-%m-%d")
            else:
                period_label = str(hdr)
        else:
            period_label = get_column_letter(col_idx + 1)

        if any(not isinstance(v, (int, float)) for v in [a, l, e] if v is not None):
            period_results.append(f"  {period_label}: SKIP (non-numeric values)")
            continue

        a_val = float(a) if a is not None else 0.0
        l_val = float(l) if l is not None else 0.0
        e_val = float(e) if e is not None else 0.0

        diff = a_val - (l_val + e_val)
        tolerance = 0.01  # allow for floating point rounding
        if abs(diff) <= tolerance:
            period_results.append(f"  {period_label}: PASS  (Assets={a_val:,.2f} = L+E={l_val+e_val:,.2f})")
        else:
            all_pass = False
            period_results.append(
                f"  {period_label}: FAIL  (Assets={a_val:,.2f} ≠ L+E={l_val+e_val:,.2f}, diff={diff:+,.2f})"
            )

    overall = "BALANCED" if all_pass else "IMBALANCED"
    lines = [f"Balance Check: {overall}", ""] + period_results

    meta = {
        "command": "validate-balance",
        "path": path,
        "sheet": sheet,
        "result": overall,
    }
    return format_output(meta, "\n".join(lines))
