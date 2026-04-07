from pathlib import Path
from openpyxl.utils import get_column_letter
from excel_explorer.workbook import load_workbook, get_sheet_names, get_named_ranges
from excel_explorer.formatters import format_output


def overview(path: str) -> str:
    """Return workbook overview: sheets, dimensions, formula counts, named ranges."""
    wb = load_workbook(path)
    fname = Path(path).name
    sheets = get_sheet_names(wb)
    nr = get_named_ranges(wb)

    lines = []
    lines.append("sheets:")
    for sn in sheets:
        ws = wb[sn]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        formula_count = 0
        scan_rows = min(max_row, 10000)
        scan_cols = min(max_col, 100)
        for row in ws.iter_rows(min_row=1, max_row=scan_rows, max_col=scan_cols):
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

        lines.append(f"  - name: {sn}")
        lines.append(f"    rows: {max_row}")
        lines.append(f"    cols: {max_col}")
        lines.append(f"    formulas: {formula_count}")

    if nr:
        lines.append("")
        lines.append("named_ranges:")
        for name, target in nr.items():
            lines.append(f"  - {name} -> {target}")

    meta = {
        "file": fname,
        "sheets": len(sheets),
        "named_ranges": len(nr),
    }
    return format_output(meta, "\n".join(lines))


def describe(path: str, sheet: str, limit: int = 50, offset: int = 0,
             max_cols: int = 20, col_offset: int = 0) -> str:
    """Describe a sheet: headers, dimensions, merged cells, data regions."""
    wb = load_workbook(path)
    ws = wb[sheet]
    fname = Path(path).name
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    lines = []
    lines.append(f"rows: {max_row}")
    lines.append(f"cols: {max_col}")

    merged = list(ws.merged_cells.ranges)
    if merged:
        lines.append(f"merged_cells: {len(merged)}")
        for m in merged[:10]:
            lines.append(f"  - {m}")

    lines.append("")
    lines.append("header_row:")
    first_row = list(ws.iter_rows(min_row=1, max_row=1, max_col=min(max_col, max_cols + col_offset)))[0]
    for cell in first_row[col_offset:col_offset + max_cols]:
        if cell.value is not None:
            lines.append(f"  {cell.coordinate}: {cell.value}")

    lines.append("")
    lines.append("data_preview:")
    start_row = 1 + offset
    end_row = min(max_row, start_row + limit - 1)
    start_col = 1 + col_offset
    end_col = min(max_col, start_col + max_cols - 1)

    for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                            min_col=start_col, max_col=end_col):
        vals = []
        for cell in row:
            v = cell.value
            if v is None:
                vals.append("")
            elif isinstance(v, str) and v.startswith("="):
                vals.append(f"[formula] {v}")
            else:
                vals.append(str(v))
        lines.append(f"  {row[0].coordinate}: {' | '.join(vals)}")

    truncated = end_row < max_row or end_col < max_col
    col_start_letter = get_column_letter(start_col)
    col_end_letter = get_column_letter(end_col)

    meta = {
        "file": fname,
        "sheet": sheet,
        "showing": f"rows {start_row}-{end_row} of {max_row}, cols {col_start_letter}-{col_end_letter} of {get_column_letter(max_col)}",
        "truncated": truncated,
    }
    return format_output(meta, "\n".join(lines))


def named_ranges(path: str) -> str:
    """List all named ranges with their targets."""
    wb = load_workbook(path)
    fname = Path(path).name
    nr = get_named_ranges(wb)

    if not nr:
        meta = {"file": fname, "named_ranges": 0}
        return format_output(meta, "No named ranges defined.")

    lines = []
    for name, target in nr.items():
        lines.append(f"  - {name}: {target}")

    meta = {"file": fname, "named_ranges": len(nr)}
    return format_output(meta, "\n".join(lines))


def read_range(path: str, sheet: str, range_str: str, formulas: bool = False,
               limit: int = 50, offset: int = 0,
               max_cols: int = 20, col_offset: int = 0) -> str:
    """Read a range of cells. If formulas=True, show both value and formula."""
    wb_formulas = load_workbook(path, data_only=False)
    wb_values = load_workbook(path, data_only=True) if formulas else None
    ws_f = wb_formulas[sheet]
    ws_v = wb_values[sheet] if wb_values else None
    fname = Path(path).name

    cells = ws_f[range_str]
    if not isinstance(cells, tuple):
        cells = ((cells,),)
    elif cells and not isinstance(cells[0], tuple):
        cells = (cells,)

    rows = list(cells)
    total_rows = len(rows)
    rows = rows[offset:offset + limit]

    lines = []
    for row in rows:
        row_cells = list(row)[col_offset:col_offset + max_cols]
        for cell in row_cells:
            val = cell.value
            if formulas and ws_v:
                computed = ws_v[cell.coordinate].value
                if isinstance(val, str) and val.startswith("="):
                    lines.append(f"  {cell.coordinate}: {computed}  (formula: {val})")
                else:
                    lines.append(f"  {cell.coordinate}: {val}")
            else:
                lines.append(f"  {cell.coordinate}: {val}")

    showing_end = min(offset + limit, total_rows)
    meta = {
        "file": fname,
        "sheet": sheet,
        "range": range_str,
        "showing": f"rows {offset + 1}-{showing_end} of {total_rows}",
        "truncated": showing_end < total_rows,
    }
    return format_output(meta, "\n".join(lines))


def read_row(path: str, sheet: str, row: int, max_cols: int = 20, col_offset: int = 0,
             formulas: bool = False) -> str:
    """Read a full row with auto-detected headers."""
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]
    fname = Path(path).name
    max_col = ws.max_column or 1

    start_col = 1 + col_offset
    end_col = min(max_col, start_col + max_cols - 1)

    headers = {}
    for cell in ws.iter_cols(min_row=1, max_row=1, min_col=start_col, max_col=end_col):
        for c in cell:
            if c.value is not None:
                headers[c.column] = str(c.value)

    lines = []
    row_label = None
    for cell in ws.iter_cols(min_row=row, max_row=row, min_col=start_col, max_col=end_col):
        for c in cell:
            header = headers.get(c.column, get_column_letter(c.column))
            val = c.value
            if c.column == start_col:
                row_label = val
            if formulas and isinstance(val, str) and val.startswith("="):
                lines.append(f"  {c.coordinate} ({header}): [formula] {val}")
            else:
                lines.append(f"  {c.coordinate} ({header}): {val}")

    col_start_letter = get_column_letter(start_col)
    col_end_letter = get_column_letter(end_col)
    meta = {
        "file": fname,
        "sheet": sheet,
        "row": row,
        "row_label": row_label if row_label else "",
        "showing": f"cols {col_start_letter}-{col_end_letter} of {get_column_letter(max_col)}",
        "truncated": end_col < max_col,
    }
    return format_output(meta, "\n".join(lines))


def read_col(path: str, sheet: str, col: str, limit: int = 50, offset: int = 0,
             formulas: bool = False) -> str:
    """Read a full column."""
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]
    fname = Path(path).name
    max_row = ws.max_row or 1

    start_row = 1 + offset
    end_row = min(max_row, start_row + limit - 1)

    lines = []
    for row in ws.iter_rows(min_row=start_row, max_row=end_row):
        cell = ws[f"{col}{row[0].row}"]
        val = cell.value
        if formulas and isinstance(val, str) and val.startswith("="):
            lines.append(f"  {cell.coordinate}: [formula] {val}")
        else:
            lines.append(f"  {cell.coordinate}: {val}")

    meta = {
        "file": fname,
        "sheet": sheet,
        "column": col,
        "showing": f"rows {start_row}-{end_row} of {max_row}",
        "truncated": end_row < max_row,
    }
    return format_output(meta, "\n".join(lines))


def detect_tables(path: str, sheet: str, limit: int = 50, offset: int = 0) -> str:
    """Auto-detect contiguous table regions on a sheet."""
    wb = load_workbook(path)
    ws = wb[sheet]
    fname = Path(path).name
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    tables = []
    current_start = None
    current_cols = set()
    prev_empty = True

    for row_idx in range(1, max_row + 1):
        row_has_data = False
        row_cols = set()
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                row_has_data = True
                row_cols.add(col_idx)

        if row_has_data and prev_empty:
            current_start = row_idx
            current_cols = row_cols
        elif row_has_data and not prev_empty:
            current_cols |= row_cols
        elif not row_has_data and not prev_empty and current_start is not None:
            min_c = min(current_cols) if current_cols else 1
            max_c = max(current_cols) if current_cols else 1
            headers = []
            for c in range(min_c, max_c + 1):
                v = ws.cell(row=current_start, column=c).value
                if v is not None:
                    headers.append(str(v))
            tables.append({
                "range": f"{get_column_letter(min_c)}{current_start}:{get_column_letter(max_c)}{row_idx - 1}",
                "rows": row_idx - 1 - current_start + 1,
                "cols": max_c - min_c + 1,
                "headers": headers[:10],
            })
            current_start = None

        prev_empty = not row_has_data

    # Handle table extending to last row
    if current_start is not None and current_cols:
        min_c = min(current_cols)
        max_c = max(current_cols)
        headers = []
        for c in range(min_c, max_c + 1):
            v = ws.cell(row=current_start, column=c).value
            if v is not None:
                headers.append(str(v))
        tables.append({
            "range": f"{get_column_letter(min_c)}{current_start}:{get_column_letter(max_c)}{max_row}",
            "rows": max_row - current_start + 1,
            "cols": max_c - min_c + 1,
            "headers": headers[:10],
        })

    total = len(tables)
    tables = tables[offset:offset + limit]

    lines = []
    for i, t in enumerate(tables):
        lines.append(f"  table_{i + offset + 1}:")
        lines.append(f"    range: {t['range']}")
        lines.append(f"    rows: {t['rows']}")
        lines.append(f"    cols: {t['cols']}")
        lines.append(f"    headers: {t['headers']}")

    meta = {
        "file": fname,
        "sheet": sheet,
        "tables_found": total,
        "showing": f"{offset + 1}-{offset + len(tables)} of {total}",
        "truncated": offset + len(tables) < total,
    }
    return format_output(meta, "\n".join(lines))
