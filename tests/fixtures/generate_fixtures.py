"""Generate synthetic test workbooks for integration testing.

All data is completely fabricated. No real business names, locations, or financials.
Run this script to regenerate the fixture .xlsx files in this directory.
"""
import datetime
import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side

FIXTURES_DIR = Path(__file__).parent
random.seed(42)  # Reproducible


def _rand_amount(low=1000, high=50000):
    return round(random.uniform(low, high), 2)


def create_balance_sheet_pl():
    """Mimics a QBO-exported financial statement: values only, no formulas.

    Two sheets: Balance Sheet and Income Statement, monthly columns, standard
    accounting line items.
    """
    wb = Workbook()
    months = [datetime.datetime(2025, m, 1) for m in range(1, 13)]

    # --- Balance Sheet ---
    ws = wb.active
    ws.title = "Balance Sheet"
    ws["A1"] = "Northwind Traders"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Balance Sheet Jan 2025 - Dec 2025"
    ws["A3"] = "Portland"

    # Date headers
    for i, dt in enumerate(months):
        ws.cell(row=5, column=i + 2, value=dt)
        ws.cell(row=5, column=i + 2).font = Font(bold=True)

    rows = [
        ("Assets", None),
        ("Current Assets", None),
        ("Cash", lambda: _rand_amount(20000, 80000)),
        ("Accounts Receivable", lambda: _rand_amount(5000, 25000)),
        ("Inventory", lambda: _rand_amount(8000, 15000)),
        ("Prepaid Expenses", lambda: _rand_amount(1000, 5000)),
        ("Total Current Assets", None),
        ("", None),
        ("Fixed Assets", None),
        ("Property & Equipment", lambda: _rand_amount(100000, 150000)),
        ("Less: Accumulated Depreciation", lambda: _rand_amount(-50000, -20000)),
        ("Total Fixed Assets", None),
        ("", None),
        ("Total Assets", None),
        ("", None),
        ("Liabilities", None),
        ("Current Liabilities", None),
        ("Accounts Payable", lambda: _rand_amount(5000, 20000)),
        ("Accrued Expenses", lambda: _rand_amount(2000, 8000)),
        ("Short-Term Debt", lambda: _rand_amount(5000, 15000)),
        ("Total Current Liabilities", None),
        ("", None),
        ("Long-Term Liabilities", None),
        ("Long-Term Debt", lambda: _rand_amount(30000, 60000)),
        ("Total Long-Term Liabilities", None),
        ("", None),
        ("Total Liabilities", None),
        ("", None),
        ("Equity", None),
        ("Owner's Equity", lambda: _rand_amount(40000, 80000)),
        ("Retained Earnings", lambda: _rand_amount(10000, 30000)),
        ("Total Equity", None),
        ("", None),
        ("Total Liabilities & Equity", None),
    ]

    for r_idx, (label, gen) in enumerate(rows, start=6):
        ws.cell(row=r_idx, column=1, value=label)
        if label.startswith("Total") or label in ("Assets", "Liabilities", "Equity",
                                                    "Current Assets", "Fixed Assets",
                                                    "Current Liabilities", "Long-Term Liabilities"):
            ws.cell(row=r_idx, column=1).font = Font(bold=True)
        if gen:
            for c_idx in range(len(months)):
                ws.cell(row=r_idx, column=c_idx + 2, value=gen())

    # Fill in totals (hardcoded sums for simplicity — this mimics QBO export)
    # Current Assets total = Cash + AR + Inventory + Prepaid
    for c in range(2, 14):
        ca_total = sum(ws.cell(row=r, column=c).value or 0 for r in [8, 9, 10, 11])
        ws.cell(row=12, column=c, value=round(ca_total, 2))
        # Fixed Assets total
        fa_total = sum(ws.cell(row=r, column=c).value or 0 for r in [15, 16])
        ws.cell(row=17, column=c, value=round(fa_total, 2))
        # Total Assets
        ws.cell(row=19, column=c, value=round(ca_total + fa_total, 2))
        # Current Liabilities total
        cl_total = sum(ws.cell(row=r, column=c).value or 0 for r in [23, 24, 25])
        ws.cell(row=26, column=c, value=round(cl_total, 2))
        # Long-Term total
        lt_total = ws.cell(row=29, column=c).value or 0
        ws.cell(row=30, column=c, value=round(lt_total, 2))
        # Total Liabilities
        ws.cell(row=32, column=c, value=round(cl_total + lt_total, 2))
        # Total Equity
        eq_total = sum(ws.cell(row=r, column=c).value or 0 for r in [35, 36])
        ws.cell(row=37, column=c, value=round(eq_total, 2))
        # Total L&E = Total Liabilities + Total Equity
        ws.cell(row=39, column=c, value=round(cl_total + lt_total + eq_total, 2))

    # --- Income Statement ---
    ws2 = wb.create_sheet("Income Statement")
    ws2["A1"] = "Northwind Traders"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A2"] = "Income Statement Jan 2025 - Dec 2025"
    ws2["A3"] = "Portland"

    for i, dt in enumerate(months):
        ws2.cell(row=5, column=i + 2, value=dt)
        ws2.cell(row=5, column=i + 2).font = Font(bold=True)

    is_rows = [
        ("Gross Sales", None),
        ("Sales-Food", lambda: _rand_amount(15000, 45000)),
        ("Sales-Beverage", lambda: _rand_amount(30000, 90000)),
        ("Sales-Merchandise", lambda: _rand_amount(500, 3000)),
        ("Total Gross Sales", None),
        ("", None),
        ("Discounts & Comps", lambda: _rand_amount(-2000, -500)),
        ("Net Sales", None),
        ("", None),
        ("Cost of Goods Sold", None),
        ("COGS-Food", lambda: _rand_amount(5000, 15000)),
        ("COGS-Beverage", lambda: _rand_amount(8000, 25000)),
        ("COGS-Merchandise", lambda: _rand_amount(100, 800)),
        ("Total COGS", None),
        ("", None),
        ("Gross Profit", None),
        ("", None),
        ("Operating Expenses", None),
        ("Payroll", lambda: _rand_amount(20000, 40000)),
        ("Rent", lambda: 8500.00),
        ("Utilities", lambda: _rand_amount(1500, 4000)),
        ("Insurance", lambda: 2200.00),
        ("Marketing", lambda: _rand_amount(500, 3000)),
        ("Repairs & Maintenance", lambda: _rand_amount(200, 2000)),
        ("Supplies", lambda: _rand_amount(300, 1500)),
        ("Professional Fees", lambda: _rand_amount(500, 2000)),
        ("Total Operating Expenses", None),
        ("", None),
        ("Operating Income", None),
        ("", None),
        ("Other Income/Expense", None),
        ("Interest Expense", lambda: _rand_amount(-800, -200)),
        ("Other Income", lambda: _rand_amount(0, 500)),
        ("Total Other", None),
        ("", None),
        ("Net Income", None),
    ]

    for r_idx, (label, gen) in enumerate(is_rows, start=6):
        ws2.cell(row=r_idx, column=1, value=label)
        if label.startswith("Total") or label in ("Gross Sales", "Net Sales",
                                                    "Cost of Goods Sold", "Gross Profit",
                                                    "Operating Expenses", "Operating Income",
                                                    "Net Income", "Other Income/Expense"):
            ws2.cell(row=r_idx, column=1).font = Font(bold=True)
        if gen:
            for c_idx in range(len(months)):
                ws2.cell(row=r_idx, column=c_idx + 2, value=gen())

    # Compute totals
    for c in range(2, 14):
        # Total Gross Sales
        gs = sum(ws2.cell(row=r, column=c).value or 0 for r in [7, 8, 9])
        ws2.cell(row=10, column=c, value=round(gs, 2))
        # Net Sales
        disc = ws2.cell(row=12, column=c).value or 0
        ws2.cell(row=13, column=c, value=round(gs + disc, 2))
        # Total COGS
        cogs = sum(ws2.cell(row=r, column=c).value or 0 for r in [16, 17, 18])
        ws2.cell(row=19, column=c, value=round(cogs, 2))
        # Gross Profit
        ws2.cell(row=21, column=c, value=round(gs + disc - cogs, 2))
        # Total OpEx
        opex = sum(ws2.cell(row=r, column=c).value or 0 for r in range(24, 32))
        ws2.cell(row=32, column=c, value=round(opex, 2))
        # Operating Income
        ws2.cell(row=34, column=c, value=round(gs + disc - cogs - opex, 2))
        # Total Other
        other = sum(ws2.cell(row=r, column=c).value or 0 for r in [37, 38])
        ws2.cell(row=39, column=c, value=round(other, 2))
        # Net Income
        ws2.cell(row=41, column=c, value=round(gs + disc - cogs - opex + other, 2))

    wb.save(FIXTURES_DIR / "financial_statements.xlsx")
    print("Created financial_statements.xlsx")


def create_financial_model():
    """A financial model with formulas, cross-sheet refs, named ranges, and assumptions.

    Sheets: Assumptions, Revenue Build, P&L, Balance Sheet
    Exercises: formulas, cross-sheet refs, named ranges, SUM/IF patterns.
    """
    wb = Workbook()
    months = [datetime.datetime(2025, m, 1) for m in range(1, 13)]

    # --- Assumptions ---
    ws_a = wb.active
    ws_a.title = "Assumptions"
    ws_a["A1"] = "Model Assumptions"
    ws_a["A1"].font = Font(bold=True, size=14)
    ws_a["A2"].font = Font(bold=True)

    assumptions = [
        ("Parameter", "Value", "Unit"),
        ("Average Check Size", 42.50, "USD"),
        ("Covers Per Day - Weekday", 85, "covers"),
        ("Covers Per Day - Weekend", 140, "covers"),
        ("Food Cost %", 0.32, "%"),
        ("Beverage Cost %", 0.22, "%"),
        ("Labor Cost %", 0.28, "%"),
        ("Monthly Rent", 8500, "USD"),
        ("Annual Growth Rate", 0.05, "%"),
        ("Tax Rate", 0.21, "%"),
        ("Tip Rate", 0.18, "%"),
        ("Weekend Days Per Month", 8, "days"),
        ("Weekday Days Per Month", 22, "days"),
    ]
    for r, (param, val, unit) in enumerate(assumptions, start=3):
        ws_a.cell(row=r, column=1, value=param)
        ws_a.cell(row=r, column=2, value=val)
        ws_a.cell(row=r, column=3, value=unit)
        if r == 3:
            for c in range(1, 4):
                ws_a.cell(row=r, column=c).font = Font(bold=True)
        # Blue fill for input cells
        if val is not None and r > 3:
            ws_a.cell(row=r, column=2).fill = PatternFill("solid", fgColor="DCE6F1")

    # --- Revenue Build ---
    ws_r = wb.create_sheet("Revenue Build")
    ws_r["A1"] = "Monthly Revenue Build"
    ws_r["A1"].font = Font(bold=True, size=14)

    ws_r["A3"] = "Metric"
    ws_r["A3"].font = Font(bold=True)
    for i, dt in enumerate(months):
        ws_r.cell(row=3, column=i + 2, value=dt)
        ws_r.cell(row=3, column=i + 2).font = Font(bold=True)

    # Row 4: Weekday Covers = Covers Per Day Weekday * Weekday Days
    ws_r["A4"] = "Weekday Covers"
    for i in range(12):
        ws_r.cell(row=4, column=i + 2, value="=Assumptions!B6*Assumptions!B15")

    # Row 5: Weekend Covers = Covers Per Day Weekend * Weekend Days
    ws_r["A5"] = "Weekend Covers"
    for i in range(12):
        ws_r.cell(row=5, column=i + 2, value="=Assumptions!B7*Assumptions!B14")

    # Row 6: Total Covers
    ws_r["A6"] = "Total Covers"
    ws_r["A6"].font = Font(bold=True)
    for i in range(12):
        col = chr(66 + i)  # B, C, D, ...
        ws_r.cell(row=6, column=i + 2, value=f"={col}4+{col}5")

    # Row 7: blank
    # Row 8: Revenue = Total Covers * Avg Check
    ws_r["A8"] = "Gross Revenue"
    ws_r["A8"].font = Font(bold=True)
    for i in range(12):
        col = chr(66 + i)
        ws_r.cell(row=8, column=i + 2, value=f"={col}6*Assumptions!B4")

    # Row 9: Food Revenue (60% of gross)
    ws_r["A9"] = "Food Revenue"
    for i in range(12):
        col = chr(66 + i)
        ws_r.cell(row=9, column=i + 2, value=f"={col}8*0.6")

    # Row 10: Beverage Revenue (40% of gross)
    ws_r["A10"] = "Beverage Revenue"
    for i in range(12):
        col = chr(66 + i)
        ws_r.cell(row=10, column=i + 2, value=f"={col}8*0.4")

    # --- P&L ---
    ws_pl = wb.create_sheet("P&L")
    ws_pl["A1"] = "Profit & Loss Statement"
    ws_pl["A1"].font = Font(bold=True, size=14)

    ws_pl["A3"] = "Line Item"
    ws_pl["A3"].font = Font(bold=True)
    for i, dt in enumerate(months):
        ws_pl.cell(row=3, column=i + 2, value=dt)
        ws_pl.cell(row=3, column=i + 2).font = Font(bold=True)

    pl_items = [
        ("Gross Revenue", "='Revenue Build'!{col}8", True),
        ("Food Revenue", "='Revenue Build'!{col}9", False),
        ("Beverage Revenue", "='Revenue Build'!{col}10", False),
        ("", None, False),
        ("Cost of Goods Sold", None, True),
        ("Food COGS", "={col}5*Assumptions!B8", False),
        ("Beverage COGS", "={col}6*Assumptions!B9", False),
        ("Total COGS", "={col}9+{col}10", True),
        ("", None, False),
        ("Gross Profit", "={col}4-{col}11", True),
        ("Gross Margin %", "={col}13/{col}4", False),
        ("", None, False),
        ("Operating Expenses", None, True),
        ("Labor", "={col}4*Assumptions!B10", False),
        ("Rent", "=Assumptions!B11", False),
        ("Utilities", None, False),  # hardcoded
        ("Marketing", None, False),  # hardcoded
        ("Other OpEx", None, False),  # hardcoded
        ("Total Operating Expenses", "=SUM({col}17:{col}21)", True),
        ("", None, False),
        ("Operating Income", "={col}13-{col}22", True),
        ("Tax", "={col}23*Assumptions!B12", False),
        ("", None, False),
        ("Net Income", "={col}23-{col}24", True),
    ]

    for r_offset, (label, formula_tpl, bold) in enumerate(pl_items):
        r = r_offset + 4
        ws_pl.cell(row=r, column=1, value=label)
        if bold:
            ws_pl.cell(row=r, column=1).font = Font(bold=True)
            ws_pl.cell(row=r, column=1).border = Border(
                bottom=Side(style="thin"))
        if formula_tpl:
            for i in range(12):
                col = chr(66 + i)
                ws_pl.cell(row=r, column=i + 2, value=formula_tpl.format(col=col))
        elif label in ("Utilities",):
            for i in range(12):
                ws_pl.cell(row=r, column=i + 2, value=_rand_amount(1500, 3500))
        elif label in ("Marketing",):
            for i in range(12):
                ws_pl.cell(row=r, column=i + 2, value=_rand_amount(800, 2500))
        elif label in ("Other OpEx",):
            for i in range(12):
                ws_pl.cell(row=r, column=i + 2, value=_rand_amount(300, 1200))
            # Introduce an anomaly: one month is a formula instead of hardcoded
            ws_pl.cell(row=r, column=8, value="=G21*1.1")  # col H = anomaly

    # --- Balance Sheet (simple) ---
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs["A1"] = "Balance Sheet"
    ws_bs["A1"].font = Font(bold=True, size=14)

    ws_bs["A3"] = ""
    for i, dt in enumerate(months):
        ws_bs.cell(row=3, column=i + 2, value=dt)

    bs_labels = [
        "Total Assets", "Total Liabilities", "Total Equity"
    ]
    for r_offset, label in enumerate(bs_labels):
        r = r_offset + 4
        ws_bs.cell(row=r, column=1, value=label)
        ws_bs.cell(row=r, column=1).font = Font(bold=True)
        for i in range(12):
            if label == "Total Assets":
                ws_bs.cell(row=r, column=i + 2, value=_rand_amount(150000, 250000))
            elif label == "Total Liabilities":
                assets = ws_bs.cell(row=4, column=i + 2).value
                equity = _rand_amount(50000, 100000)
                ws_bs.cell(row=r, column=i + 2, value=round(assets - equity, 2))
                ws_bs.cell(row=6, column=i + 2, value=round(equity, 2))

    # Named ranges
    from openpyxl.workbook.defined_name import DefinedName
    wb.defined_names.add(DefinedName(
        name="GrowthRate", attr_text="Assumptions!$B$12"))
    wb.defined_names.add(DefinedName(
        name="TaxRate", attr_text="Assumptions!$B$13"))
    wb.defined_names.add(DefinedName(
        name="AvgCheck", attr_text="Assumptions!$B$4"))

    wb.save(FIXTURES_DIR / "financial_model.xlsx")
    print("Created financial_model.xlsx")


def create_sales_analysis():
    """A workbook with SUMPRODUCT, named ranges, and a data sheet.

    Mimics the Peer Group Analysis pattern: a large data sheet with named range
    columns, and analysis sheets that use SUMPRODUCT to aggregate.
    """
    wb = Workbook()
    random.seed(42)

    # --- Sales Data sheet ---
    ws_data = wb.active
    ws_data.title = "Sales Data"

    headers = ["Date", "Region", "Store", "Category", "Zip", "Units", "Revenue"]
    for c, h in enumerate(headers, 1):
        ws_data.cell(row=1, column=c, value=h)
        ws_data.cell(row=1, column=c).font = Font(bold=True)

    regions = ["West", "East", "Central"]
    stores = [
        ("Maple & 3rd", "West", "97201"),
        ("Harbor View", "West", "97209"),
        ("Eastside Market", "East", "97214"),
        ("Central Station", "Central", "97232"),
        ("River District", "West", "97227"),
        ("Pearl Commons", "Central", "97209"),
    ]
    categories = ["Food", "Beverage", "Merchandise"]

    row = 2
    for month in range(1, 25):  # 24 months of data
        dt = datetime.datetime(2023 + (month - 1) // 12, ((month - 1) % 12) + 1, 15)
        for store_name, region, zip_code in stores:
            for cat in categories:
                units = random.randint(50, 500)
                rev = round(units * random.uniform(8, 45), 2)
                ws_data.cell(row=row, column=1, value=dt)
                ws_data.cell(row=row, column=2, value=region)
                ws_data.cell(row=row, column=3, value=store_name)
                ws_data.cell(row=row, column=4, value=cat)
                ws_data.cell(row=row, column=5, value=zip_code)
                ws_data.cell(row=row, column=6, value=units)
                ws_data.cell(row=row, column=7, value=rev)
                row += 1

    max_data_row = row - 1

    # Named ranges
    from openpyxl.workbook.defined_name import DefinedName
    for col_idx, name in [(1, "SaleDate"), (2, "Region"), (3, "Store"),
                           (4, "Category"), (5, "SaleZip"), (6, "Units"),
                           (7, "SaleRevenue")]:
        col_letter = chr(64 + col_idx)
        wb.defined_names.add(DefinedName(
            name=name,
            attr_text=f"'Sales Data'!${col_letter}$2:${col_letter}${max_data_row}"))

    # --- Store Analysis sheet ---
    ws_store = wb.create_sheet("Store Analysis")
    ws_store["A1"] = "Revenue by Store"
    ws_store["A1"].font = Font(bold=True, size=14)

    ws_store["A3"] = "Store"
    ws_store["B3"] = "Total Revenue"
    ws_store["C3"] = "Total Units"
    ws_store["D3"] = "Avg Revenue/Unit"
    for c in range(1, 5):
        ws_store.cell(row=3, column=c).font = Font(bold=True)

    for i, (store_name, _, _) in enumerate(stores):
        r = i + 4
        ws_store.cell(row=r, column=1, value=store_name)
        ws_store.cell(row=r, column=2,
                      value=f'=SUMPRODUCT(SaleRevenue*(Store=A{r}))')
        ws_store.cell(row=r, column=3,
                      value=f'=SUMPRODUCT(Units*(Store=A{r}))')
        ws_store.cell(row=r, column=4,
                      value=f'=IF(C{r}>0,B{r}/C{r},0)')

    # Total row
    r = len(stores) + 4
    ws_store.cell(row=r, column=1, value="Total")
    ws_store.cell(row=r, column=1).font = Font(bold=True)
    ws_store.cell(row=r, column=2, value=f"=SUM(B4:B{r-1})")
    ws_store.cell(row=r, column=3, value=f"=SUM(C4:C{r-1})")
    ws_store.cell(row=r, column=4, value=f"=IF(C{r}>0,B{r}/C{r},0)")

    # --- Region Analysis sheet ---
    ws_region = wb.create_sheet("Region Analysis")
    ws_region["A1"] = "Revenue by Region"
    ws_region["A1"].font = Font(bold=True, size=14)

    ws_region["A3"] = "Region"
    ws_region["B3"] = "Total Revenue"
    ws_region["C3"] = "Store Count"
    for c in range(1, 4):
        ws_region.cell(row=3, column=c).font = Font(bold=True)

    for i, region in enumerate(regions):
        r = i + 4
        ws_region.cell(row=r, column=1, value=region)
        ws_region.cell(row=r, column=2,
                       value=f'=SUMPRODUCT(SaleRevenue*(Region=A{r}))')
        store_count = sum(1 for _, reg, _ in stores if reg == region)
        ws_region.cell(row=r, column=3, value=store_count)

    # --- Monthly Trend sheet ---
    ws_trend = wb.create_sheet("Monthly Trend")
    ws_trend["A1"] = "Monthly Revenue Trend"
    ws_trend["A1"].font = Font(bold=True, size=14)

    ws_trend["A3"] = "Month"
    ws_trend["B3"] = "Revenue"
    ws_trend["C3"] = "MoM Change"
    ws_trend["D3"] = "MoM Growth %"
    for c in range(1, 5):
        ws_trend.cell(row=3, column=c).font = Font(bold=True)

    for m in range(24):
        r = m + 4
        dt = datetime.datetime(2023 + m // 12, (m % 12) + 1, 1)
        ws_trend.cell(row=r, column=1, value=dt)
        ws_trend.cell(row=r, column=2,
                      value=f'=SUMPRODUCT(SaleRevenue*(SaleDate=A{r}))')
        if m > 0:
            ws_trend.cell(row=r, column=3, value=f"=B{r}-B{r-1}")
            ws_trend.cell(row=r, column=4, value=f"=IF(B{r-1}<>0,C{r}/B{r-1},0)")

    wb.save(FIXTURES_DIR / "sales_analysis.xlsx")
    print("Created sales_analysis.xlsx")


def create_opex_detail():
    """Transaction-level operating expense data with SUMPRODUCT/COUNTIFS summaries.

    Mimics the OpEx workbook pattern: raw transaction data with vendor parsing
    formulas and summary aggregations.
    """
    wb = Workbook()
    random.seed(123)

    vendors = [
        "V:Sysco B:44829103", "V:USFoods B:77281934", "V:PepsiCo B:33910284",
        "V:CleanCorp B:55192837", "V:RepairPro B:66738291", "V:InsureCo B:99281734",
        "V:PowerGrid B:22918374", "V:WaterWorks B:11827364", "V:OfficeMax B:88372619",
        "V:MarketingCo B:44738291", "Pay date 1/15", "Pay date 2/1", "Pay date 2/15",
    ]

    categories = {
        "V:Sysco": "Food Cost",
        "V:USFoods": "Food Cost",
        "V:PepsiCo": "Beverage Cost",
        "V:CleanCorp": "Cleaning",
        "V:RepairPro": "Repairs",
        "V:InsureCo": "Insurance",
        "V:PowerGrid": "Utilities",
        "V:WaterWorks": "Utilities",
        "V:OfficeMax": "Office Supplies",
        "V:MarketingCo": "Marketing",
        "Pay date": "Payroll",
    }

    # --- Transactions sheet ---
    ws = wb.active
    ws.title = "Transactions"
    ws["A1"] = "Northwind Traders"
    ws["A2"] = "Operating Expenses Detail"
    ws["A3"] = "Portland"

    headers = ["Date", "Reference", "Description", "Amount", "Vendor", "Category"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=5, column=c, value=h)
        ws.cell(row=5, column=c).font = Font(bold=True)

    row = 6
    for month in range(1, 7):
        dt_base = datetime.datetime(2025, month, 1)
        num_txns = random.randint(30, 60)
        for _ in range(num_txns):
            day = random.randint(1, 28)
            dt = datetime.datetime(2025, month, day)
            vendor_desc = random.choice(vendors)
            amount = round(random.uniform(50, 5000), 2)
            if "Pay date" in vendor_desc:
                amount = round(random.uniform(8000, 15000), 2)

            ws.cell(row=row, column=1, value=dt)
            ws.cell(row=row, column=2, value=f"T{dt.strftime('%Y-%m-%d')}{random.randint(1000, 9999)}")
            ws.cell(row=row, column=3, value=vendor_desc)
            ws.cell(row=row, column=4, value=amount)
            # Vendor extraction formula
            ws.cell(row=row, column=5,
                    value=f'=TRIM(IFERROR(LEFT(C{row},FIND(" ",C{row})),C{row}))')
            # Category lookup (simplified)
            ws.cell(row=row, column=6,
                    value=f'=IFERROR(VLOOKUP(E{row},Categories!A:B,2,FALSE),"Other")')
            row += 1

    max_row = row - 1

    # --- Categories lookup sheet ---
    ws_cat = wb.create_sheet("Categories")
    ws_cat["A1"] = "Vendor Prefix"
    ws_cat["B1"] = "Category"
    ws_cat["A1"].font = Font(bold=True)
    ws_cat["B1"].font = Font(bold=True)
    for i, (prefix, cat) in enumerate(categories.items()):
        ws_cat.cell(row=i + 2, column=1, value=prefix)
        ws_cat.cell(row=i + 2, column=2, value=cat)

    # --- Summary sheet ---
    ws_sum = wb.create_sheet("Summary")
    ws_sum["A1"] = "Monthly Expense Summary"
    ws_sum["A1"].font = Font(bold=True, size=14)

    ws_sum["A3"] = "Category"
    months_dt = [datetime.datetime(2025, m, 1) for m in range(1, 7)]
    for i, dt in enumerate(months_dt):
        ws_sum.cell(row=3, column=i + 2, value=dt)
        ws_sum.cell(row=3, column=i + 2).font = Font(bold=True)

    unique_cats = sorted(set(categories.values()))
    for r_offset, cat in enumerate(unique_cats):
        r = r_offset + 4
        ws_sum.cell(row=r, column=1, value=cat)
        for i, dt in enumerate(months_dt):
            end_dt = datetime.datetime(2025, dt.month + 1, 1) if dt.month < 12 else datetime.datetime(2026, 1, 1)
            ws_sum.cell(row=r, column=i + 2,
                        value=f'=SUMPRODUCT(Transactions!$D$6:$D${max_row},'
                              f'(Transactions!$F$6:$F${max_row}=A{r}),'
                              f'(Transactions!$A$6:$A${max_row}>={dt.strftime("%m/%d/%Y")}),'
                              f'(Transactions!$A$6:$A${max_row}<{end_dt.strftime("%m/%d/%Y")}))')

    # Total row
    r_total = len(unique_cats) + 4
    ws_sum.cell(row=r_total, column=1, value="Total")
    ws_sum.cell(row=r_total, column=1).font = Font(bold=True)
    for i in range(6):
        col = chr(66 + i)
        ws_sum.cell(row=r_total, column=i + 2,
                    value=f"=SUM({col}4:{col}{r_total - 1})")

    # Transaction count
    r_count = r_total + 2
    ws_sum.cell(row=r_count, column=1, value="Transaction Count")
    for i, dt in enumerate(months_dt):
        end_dt = datetime.datetime(2025, dt.month + 1, 1) if dt.month < 12 else datetime.datetime(2026, 1, 1)
        ws_sum.cell(row=r_count, column=i + 2,
                    value=f'=COUNTIFS(Transactions!$A$6:$A${max_row},">="&{dt.strftime("%m/%d/%Y")},'
                          f'Transactions!$A$6:$A${max_row},"<"&{end_dt.strftime("%m/%d/%Y")})')

    wb.save(FIXTURES_DIR / "opex_detail.xlsx")
    print("Created opex_detail.xlsx")


if __name__ == "__main__":
    create_balance_sheet_pl()
    create_financial_model()
    create_sales_analysis()
    create_opex_detail()
    print("\nAll fixtures generated.")
