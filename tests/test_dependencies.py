import pytest
from excel_explorer.dependencies import (
    build_dag,
    trace_cell,
    find_dependents,
    formula_map,
    find_inputs,
    sheet_flow,
)


def test_build_dag(tmp_workbook):
    dag = build_dag(str(tmp_workbook))
    assert dag is not None
    assert hasattr(dag, "dsp")


def test_trace_cell_simple(tmp_workbook):
    result = trace_cell(str(tmp_workbook), "Summary", "B4")
    # B4 = B2 - B3, both are hardcoded values
    assert "B2" in result
    assert "B3" in result
    assert "[INPUT]" in result


def test_trace_cell_nested(tmp_workbook):
    result = trace_cell(str(tmp_workbook), "Summary", "B5")
    # B5 = B4/B2, B4 = B2-B3, so should see B2, B3 as leaves
    assert "B4" in result
    assert "B2" in result


def test_trace_cell_cross_sheet(tmp_workbook):
    result = trace_cell(str(tmp_workbook), "Projections", "B3")
    # B3 = B2*(1+Assumptions!B2), B2 = Summary!B2
    assert "Assumptions" in result or "assumptions" in result.lower()


def test_trace_cell_depth_limit(tmp_workbook):
    result = trace_cell(str(tmp_workbook), "Summary", "B5", depth=1)
    assert "depth: 1" in result


def test_find_dependents(tmp_workbook):
    result = find_dependents(str(tmp_workbook), "Summary", "B2")
    # B2 is used by B4 (=B2-B3), B5 (=B4/B2), B6 (=SUM(B2:C2)), and Projections!B2
    assert "B4" in result or "B5" in result or "B6" in result


def test_formula_map(tmp_workbook):
    result = formula_map(str(tmp_workbook), "Summary")
    assert "pattern" in result.lower() or "=" in result


def test_find_inputs(tmp_workbook):
    result = find_inputs(str(tmp_workbook))
    assert "INPUT" in result or "input" in result.lower()


def test_find_inputs_with_sheet_filter(tmp_workbook):
    result = find_inputs(str(tmp_workbook), sheet="Assumptions")
    assert "Assumptions" in result


def test_sheet_flow(tmp_workbook):
    result = sheet_flow(str(tmp_workbook))
    assert "Summary" in result
    assert "Projections" in result
    assert "->" in result


def test_sheet_flow_leading_cross_sheet_ref(tmp_path):
    """A formula that starts with a sheet ref (=Assumptions!B2) must produce an edge."""
    from openpyxl import Workbook

    path = tmp_path / "leading_ref.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    ws["B2"] = 7
    model = wb.create_sheet("Model")
    model["A1"] = "=Assumptions!B2"
    wb.save(path)

    result = sheet_flow(str(path))
    assert '"Assumptions" -> "Model"' in result


def test_trace_cell_does_not_substring_match_longer_refs(tmp_path):
    """Tracing A1 must not silently resolve to A10 when A1 is not in the graph."""
    from openpyxl import Workbook

    path = tmp_path / "anchor.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A10"] = 7
    ws["A11"] = 3
    ws["B1"] = "=A10+A11"
    wb.save(path)

    result = trace_cell(str(path), "S1", "A1")
    assert "not found" in result
    assert "A10" not in result


def test_sheet_flow_and_overview_skip_chartsheets(tmp_path):
    """Chartsheets have no cells and must not crash sheet iteration."""
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from excel_explorer.explorer import overview
    from excel_explorer.search import search

    path = tmp_path / "chart.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Revenue"
    ws["B1"] = 100
    cs = wb.create_chartsheet()
    cs.title = "Charts"
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=1))
    cs.add_chart(chart)
    wb.save(path)

    result = overview(str(path))
    assert "Charts" in result and "chartsheet" in result

    result = search(str(path), "Revenue")
    assert "A1" in result

    result = sheet_flow(str(path))
    assert "Data" in result
