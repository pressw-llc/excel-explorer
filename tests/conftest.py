import pytest
from pathlib import Path
from openpyxl import Workbook


@pytest.fixture
def tmp_workbook(tmp_path):
    """Create a simple test workbook with known structure."""
    path = tmp_path / "test.xlsx"
    wb = Workbook()

    # Sheet 1: Summary (with formulas referencing Sheet2)
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Metric"
    ws1["B1"] = "Jan"
    ws1["C1"] = "Feb"
    ws1["A2"] = "Revenue"
    ws1["B2"] = 1000
    ws1["C2"] = 1200
    ws1["A3"] = "COGS"
    ws1["B3"] = 400
    ws1["C3"] = 480
    ws1["A4"] = "Gross Profit"
    ws1["B4"] = "=B2-B3"
    ws1["C4"] = "=C2-C3"
    ws1["A5"] = "GP Margin"
    ws1["B5"] = "=B4/B2"
    ws1["C5"] = "=C4/C2"
    ws1["A6"] = "Total Revenue"
    ws1["B6"] = "=SUM(B2:C2)"

    # Sheet 2: Assumptions
    ws2 = wb.create_sheet("Assumptions")
    ws2["A1"] = "Parameter"
    ws2["B1"] = "Value"
    ws2["A2"] = "Growth Rate"
    ws2["B2"] = 0.05
    ws2["A3"] = "Tax Rate"
    ws2["B3"] = 0.21
    ws2["A4"] = "Headcount"
    ws2["B4"] = 25

    # Sheet 3: Projections (references Assumptions)
    ws3 = wb.create_sheet("Projections")
    ws3["A1"] = "Month"
    ws3["B1"] = "Revenue"
    ws3["A2"] = "Jan"
    ws3["B2"] = "=Summary!B2"
    ws3["A3"] = "Feb"
    ws3["B3"] = "=B2*(1+Assumptions!B2)"

    wb.save(path)
    return path


@pytest.fixture
def empty_workbook(tmp_path):
    """Create a workbook with no formulas (like a QBO export)."""
    path = tmp_path / "empty.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    ws["A1"] = "Assets"
    ws["B1"] = 50000
    ws["A2"] = "Liabilities"
    ws["B2"] = 20000
    ws["A3"] = "Equity"
    ws["B3"] = 30000
    wb.save(path)
    return path


@pytest.fixture
def real_workbook_dir():
    """Path to real test workbooks for integration testing.

    Set the XLX_TEST_WORKBOOKS env var to a directory containing .xlsx files.
    Integration tests are skipped when unset.
    """
    import os
    env_path = os.environ.get("XLX_TEST_WORKBOOKS")
    if not env_path:
        pytest.skip("Set XLX_TEST_WORKBOOKS env var to run integration tests")
    path = Path(env_path)
    if not path.exists():
        pytest.skip(f"Test workbook directory not found: {path}")
    return path
