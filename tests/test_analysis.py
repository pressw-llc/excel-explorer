import pytest
from openpyxl import Workbook
from excel_explorer.analysis import summarize_assumptions, compare_periods, find_anomalies, validate_balance


@pytest.fixture
def financial_workbook(tmp_path):
    """Workbook with time-series financial data for analysis testing."""
    import datetime
    path = tmp_path / "financial.xlsx"
    wb = Workbook()

    ws = wb.active
    ws.title = "Income Statement"
    ws["A1"] = ""
    ws["B1"] = datetime.datetime(2025, 1, 1)
    ws["C1"] = datetime.datetime(2025, 2, 1)
    ws["D1"] = datetime.datetime(2025, 3, 1)
    ws["A2"] = "Revenue"
    ws["B2"] = 10000
    ws["C2"] = 12000
    ws["D2"] = 11000
    ws["A3"] = "COGS"
    ws["B3"] = "=B2*0.4"
    ws["C3"] = "=C2*0.4"
    ws["D3"] = 5000  # anomaly: hardcoded override
    ws["A4"] = "Gross Profit"
    ws["B4"] = "=B2-B3"
    ws["C4"] = "=C2-C3"
    ws["D4"] = "=D2-D3"

    ws2 = wb.create_sheet("Balance Sheet")
    ws2["A1"] = ""
    ws2["B1"] = datetime.datetime(2025, 1, 31)
    ws2["A2"] = "Total Assets"
    ws2["B2"] = 100000
    ws2["A3"] = "Total Liabilities"
    ws2["B3"] = 60000
    ws2["A4"] = "Total Equity"
    ws2["B4"] = 40000

    ws3 = wb.create_sheet("Assumptions")
    ws3["A1"] = "Parameter"
    ws3["B1"] = "Value"
    ws3["A2"] = "COGS %"
    ws3["B2"] = 0.4
    ws3["A3"] = "Growth Rate"
    ws3["B3"] = 0.05

    wb.save(path)
    return path


def test_compare_periods(financial_workbook):
    result = compare_periods(str(financial_workbook), "Income Statement", 2)
    assert "Revenue" in result
    assert "change" in result.lower() or "growth" in result.lower()


def test_find_anomalies(financial_workbook):
    result = find_anomalies(str(financial_workbook), "Income Statement")
    assert "D3" in result or "anomal" in result.lower()


def test_validate_balance(financial_workbook):
    result = validate_balance(str(financial_workbook), "Balance Sheet")
    assert "pass" in result.lower() or "balanced" in result.lower() or "match" in result.lower()


def test_summarize_assumptions(financial_workbook):
    result = summarize_assumptions(str(financial_workbook))
    assert "Assumptions" in result or "assumptions" in result.lower()


def test_find_anomalies_detects_cross_column_pattern_break(tmp_path):
    """Row-wise comparison must normalize column letters: =B2+B3 vs =C2+C3 are the
    same pattern, while =E2*E3 is a structural break."""
    path = tmp_path / "pattern_break.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Model"
    for col in "BCD":
        ws[f"{col}2"] = 1
        ws[f"{col}3"] = 2
        ws[f"{col}5"] = f"={col}2+{col}3"
    ws["E2"] = 1
    ws["E3"] = 2
    ws["E5"] = "=E2*E3"
    wb.save(path)

    result = find_anomalies(str(path), "Model")
    assert "E5" in result
    assert "pattern differs" in result


def test_find_anomalies_ignores_literal_majority_rows(tmp_path):
    """Literals in a mostly-literal row are normal data, not anomalies."""
    path = tmp_path / "literal_majority.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S"
    for col in "BCDEF":
        ws[f"{col}1"] = 100
    ws["G1"] = "=B1+C1"
    wb.save(path)

    result = find_anomalies(str(path), "S")
    assert "hardcoded value" not in result
